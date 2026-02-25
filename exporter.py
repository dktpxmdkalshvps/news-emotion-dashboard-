"""
exporter.py - 분석 결과를 Excel 파일로 내보내기
openpyxl 기반 스타일링 + 요약 시트 포함
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


# ── 색상 상수 ─────────────────────────────────────────────────────────────────
COLOR = {
    "header_bg":  "1A1A2E",
    "header_fg":  "EAEAEA",
    "pos_fill":   "C8F7C5",   # 연두
    "neg_fill":   "FADBD8",   # 연빨강
    "neu_fill":   "F2F3F4",   # 연회색
    "pos_dark":   "1E8449",
    "neg_dark":   "C0392B",
    "neu_dark":   "7F8C8D",
    "row_alt":    "EBF5FB",   # 짝수행 배경
}

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


class DataExporter:
    """
    감성 분석 결과를 Excel 파일로 저장합니다.
    
    시트 구성:
      1. 📊 요약 통계    - KPI 카드 형태의 요약 + 막대/파이 차트
      2. 📰 전체 데이터  - 스타일링된 원시 데이터 테이블
      3. 👍 긍정 기사    - 긍정 기사만 필터링
      4. 👎 부정 기사    - 부정 기사만 필터링
    """

    COLUMNS = {
        "title":       "제목",
        "press":       "언론사",
        "pub_time":    "게시 시간",
        "score":       "감성 점수",
        "sentiment":   "감성",
        "matched_pos": "긍정 키워드",
        "matched_neg": "부정 키워드",
        "url":         "원문 링크",
        "crawled_at":  "수집 시각",
    }

    def __init__(self, keyword: str, output_dir: str = "output"):
        self.keyword = keyword
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export(self, df: pd.DataFrame) -> str:
        """
        DataFrame을 엑셀 파일로 저장합니다.
        
        Returns:
            저장된 파일 경로
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.keyword}_감성분석_{timestamp}.xlsx"
        path = os.path.join(self.output_dir, filename)

        # 컬럼 순서 정렬 & 한글 컬럼명 적용
        export_cols = [c for c in self.COLUMNS if c in df.columns]
        export_df = df[export_cols].rename(columns=self.COLUMNS)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # 시트 1: 전체 데이터
            export_df.to_excel(writer, sheet_name="전체 데이터", index=False)

            # 시트 2: 긍정 기사
            pos_df = export_df[export_df["감성"] == "긍정"]
            pos_df.to_excel(writer, sheet_name="긍정 기사", index=False)

            # 시트 3: 부정 기사
            neg_df = export_df[export_df["감성"] == "부정"]
            neg_df.to_excel(writer, sheet_name="부정 기사", index=False)

        # openpyxl로 다시 열어 스타일 적용
        wb = load_workbook(path)
        self._style_data_sheet(wb["전체 데이터"], df)
        self._style_data_sheet(wb["긍정 기사"],  df[df["sentiment"] == "긍정"])
        self._style_data_sheet(wb["부정 기사"],  df[df["sentiment"] == "부정"])
        self._create_summary_sheet(wb, df)

        # 시트 순서 재정렬 (요약이 맨 앞)
        wb.move_sheet("요약 통계", offset=-wb.sheetnames.index("요약 통계"))
        wb.save(path)
        return path

    # ── 데이터 시트 스타일링 ──────────────────────────────────────
    def _style_data_sheet(self, ws, df: pd.DataFrame):
        """헤더 스타일 + 행 색상 + 감성 셀 강조 + 컬럼 너비 자동 조정"""
        col_widths = {
            "제목": 55, "언론사": 16, "게시 시간": 16,
            "감성 점수": 10, "감성": 8,
            "긍정 키워드": 20, "부정 키워드": 20,
            "원문 링크": 50, "수집 시각": 18,
        }

        # ① 헤더 스타일
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=COLOR["header_bg"])
            cell.font = Font(bold=True, color=COLOR["header_fg"], size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            # 컬럼 너비
            col_letter = get_column_letter(cell.column)
            col_name = cell.value or ""
            width = col_widths.get(col_name, 15)
            ws.column_dimensions[col_letter].width = width

        ws.row_dimensions[1].height = 22

        # 감성 컬럼 인덱스 찾기
        sentiment_col = None
        for cell in ws[1]:
            if cell.value == "감성":
                sentiment_col = cell.column
                break

        # ② 데이터 행 스타일
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # 짝수행 배경
            row_fill = PatternFill("solid", fgColor=COLOR["row_alt"]) \
                       if row_idx % 2 == 0 else None

            sentiment_val = None
            if sentiment_col:
                sentiment_val = ws.cell(row_idx, sentiment_col).value

            for cell in row:
                # 기본 스타일
                cell.alignment = Alignment(
                    vertical="center", wrap_text=(cell.column == 2)
                )
                cell.border = THIN_BORDER
                if row_fill:
                    cell.fill = row_fill

                # 감성 셀 강조색 (감성 컬럼 전체 행)
                if sentiment_val == "긍정":
                    cell.fill = PatternFill("solid", fgColor=COLOR["pos_fill"])
                    if cell.column == sentiment_col:
                        cell.font = Font(bold=True, color=COLOR["pos_dark"])
                elif sentiment_val == "부정":
                    cell.fill = PatternFill("solid", fgColor=COLOR["neg_fill"])
                    if cell.column == sentiment_col:
                        cell.font = Font(bold=True, color=COLOR["neg_dark"])
                elif sentiment_val == "중립":
                    if cell.column == sentiment_col:
                        cell.font = Font(color=COLOR["neu_dark"])

        # ③ 틀 고정 (헤더 행)
        ws.freeze_panes = "A2"
        # ④ 자동 필터
        ws.auto_filter.ref = ws.dimensions

    # ── 요약 시트 생성 ────────────────────────────────────────────
    def _create_summary_sheet(self, wb, df: pd.DataFrame):
        """KPI 카드 + 차트가 포함된 요약 시트를 생성합니다."""
        ws = wb.create_sheet("요약 통계")
        ws.sheet_view.showGridLines = False

        total = len(df)
        counts = df["sentiment"].value_counts()
        pos_n = counts.get("긍정", 0)
        neg_n = counts.get("부정", 0)
        neu_n = counts.get("중립", 0)
        avg_score = round(df["score"].mean(), 3)

        # ─ 타이틀 ─
        ws.merge_cells("B2:H2")
        title_cell = ws["B2"]
        title_cell.value = f"📰 [{self.keyword}] 뉴스 감성 분석 요약 리포트"
        title_cell.font = Font(size=16, bold=True, color=COLOR["header_bg"])
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 30

        ws.merge_cells("B3:H3")
        date_cell = ws["B3"]
        date_cell.value = f"분석 일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}"
        date_cell.font = Font(size=10, color="888888")
        date_cell.alignment = Alignment(horizontal="center")

        # ─ KPI 카드 ─
        kpis = [
            ("총 기사 수",  total,      COLOR["header_bg"], "FFFFFF"),
            ("긍정 기사",   pos_n,      COLOR["pos_dark"],  "FFFFFF"),
            ("부정 기사",   neg_n,      COLOR["neg_dark"],  "FFFFFF"),
            ("중립 기사",   neu_n,      "7F8C8D",           "FFFFFF"),
            ("평균 점수",   avg_score,  "2C3E50",           "FFFFFF"),
        ]

        for i, (label, value, bg, fg) in enumerate(kpis):
            col = chr(ord("B") + i)
            label_cell = ws[f"{col}5"]
            value_cell = ws[f"{col}6"]

            label_cell.value = label
            label_cell.fill = PatternFill("solid", fgColor=bg)
            label_cell.font = Font(bold=True, color=fg, size=10)
            label_cell.alignment = Alignment(horizontal="center")
            label_cell.border = THIN_BORDER

            value_cell.value = value
            value_cell.fill = PatternFill("solid", fgColor=bg)
            value_cell.font = Font(bold=True, color=fg, size=18)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.border = THIN_BORDER
            ws.row_dimensions[6].height = 36
            ws.column_dimensions[col].width = 14

        # ─ 차트용 데이터 테이블 ─
        ws["B9"] = "감성"
        ws["C9"] = "건수"
        ws["D9"] = "비율(%)"
        for i, (label, n) in enumerate(
            [("긍정", pos_n), ("중립", neu_n), ("부정", neg_n)], start=10
        ):
            ws[f"B{i}"] = label
            ws[f"C{i}"] = n
            ws[f"D{i}"] = round(n / total * 100, 1) if total else 0

        # 테이블 헤더 스타일
        for cell in [ws["B9"], ws["C9"], ws["D9"]]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=COLOR["header_bg"])
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # ─ 막대 차트 ─
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "감성별 기사 수"
        bar_chart.y_axis.title = "건수"
        bar_chart.x_axis.title = "감성"
        bar_chart.width = 14
        bar_chart.height = 10

        data_ref = Reference(ws, min_col=3, min_row=9, max_row=12)
        cats_ref = Reference(ws, min_col=2, min_row=10, max_row=12)
        bar_chart.add_data(data_ref, titles_from_data=True)
        bar_chart.set_categories(cats_ref)
        ws.add_chart(bar_chart, "F5")

        # ─ 파이 차트 ─
        pie_chart = PieChart()
        pie_chart.title = "감성 비율"
        pie_chart.width = 14
        pie_chart.height = 10

        pie_data = Reference(ws, min_col=3, min_row=9, max_row=12)
        pie_labels = Reference(ws, min_col=2, min_row=10, max_row=12)
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_labels)
        ws.add_chart(pie_chart, "F22")
